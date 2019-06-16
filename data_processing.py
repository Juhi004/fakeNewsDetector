# Pre-processing data and feature selection
import xlrd
import openpyxl
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import nltk

# nltk.download('punkt')


vectorizerStatement = TfidfVectorizer()
vectorizerTopic = TfidfVectorizer()
vectorizerSpeaker = TfidfVectorizer()

wb11 = xlrd.open_workbook("train.xlsx")
wb21 = xlrd.open_workbook("test.xlsx")
train_dataset = wb11.sheet_by_index(0)
test_dataset = wb21.sheet_by_index(0)

workbook1 = openpyxl.load_workbook('new_train.xlsx')
workbook2 = openpyxl.load_workbook('new_test.xlsx')
train = workbook1.worksheets[0]
test = workbook2.worksheets[0]

stop_words = set(stopwords.words('english'))

text1 = []
text2 = []
text3 = []

n = train_dataset.nrows

filtered_sentence = ""

for i in range(1, n):
    word_tokens = word_tokenize(train_dataset.cell_value(i, 1))
    for w in word_tokens:
        if w not in stop_words:
            filtered_sentence = filtered_sentence + " " + w
    text1.append(filtered_sentence)

    filtered_sentence = ""
    word_tokens = word_tokenize(train_dataset.cell_value(i, 2))
    for w in word_tokens:
        if w not in stop_words:
            filtered_sentence = filtered_sentence + " " + w
    text2.append(filtered_sentence)

    filtered_sentence = ""
    word_tokens = word_tokenize(train_dataset.cell_value(i, 3))
    for w in word_tokens:
        if w not in stop_words:
            filtered_sentence = filtered_sentence + " " + w
    text3.append(filtered_sentence)

vectorizerStatement.fit(text1)
vectorizerTopic.fit(text2)
vectorizerSpeaker.fit(text3)


def encode_data(hB):

    # converting training data set

    statement = []
    topics = []
    speaker = []
    verdict = []

    filtered_sentence = ""

    n = train_dataset.nrows
    # print(n)

    print("starting encoding")

    # /////

    for i in range(0, n):
        hB.onHeartBeat()
        word_tokens = word_tokenize(train_dataset.cell_value(i, 1))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector1 = vectorizerStatement.transform([filtered_sentence])

        filtered_sentence = ""
        word_tokens = word_tokenize(train_dataset.cell_value(i, 2))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector2 = vectorizerStatement.transform([filtered_sentence])

        filtered_sentence = ""
        word_tokens = word_tokenize(train_dataset.cell_value(i, 3))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector3 = vectorizerStatement.transform([filtered_sentence])

        arr1 = vector1.toarray()
        arr2 = vector2.toarray()
        arr3 = vector3.toarray()

        tup1 = arr1.shape
        tup2 = arr2.shape
        tup3 = arr3.shape
        length1 = tup1[1]
        length2 = tup2[1]
        length3 = tup3[1]
        count1 = 0
        count2 = 0
        count3 = 0

        for j in range(0, length1):
            if arr1[0, j] != 0:
                count1 = count1 + 1
        for j in range(0, length2):
            if arr2[0, j] != 0:
                count2 = count2 + 1
        for j in range(0, length3):
            if arr3[0, j] != 0:
                count3 = count3 + 1

        sum1 = np.sum(arr1)
        sum2 = np.sum(arr2)
        sum3 = np.sum(arr3)

        if count1 != 0:
            ans1 = sum1 / count1
        else:
            ans1 = sum1
        if count2 != 0:
            ans2 = sum2 / count2
        else:
            ans2 = sum2
        if count3 != 0:
            ans3 = sum3 / count3
        else:
            ans3 = sum3

        statement.append(ans1)
        topics.append(ans2)
        speaker.append(ans3)
        verdict.append(train_dataset.cell_value(i, 0))
        hB.onHeartBeat()

    for i in range(1, n):
        j = train.cell(row=i, column=1)
        j.value = verdict[i]
    workbook1.save('new_train.xlsx')

    for i in range(1, n):
        j = train.cell(row=i, column=2)
        j.value = statement[i]
    workbook1.save('new_train.xlsx')

    for i in range(1, n):
        j = train.cell(row=i, column=3)
        j.value = topics[i]
    workbook1.save('new_train.xlsx')

    for i in range(1, n):
        j = train.cell(row=i, column=4)
        j.value = speaker[i]
    workbook1.save('new_train.xlsx')

    # converting testing data set
    statement = []
    topics = []
    speaker = []
    verdict = []

    m = test_dataset.nrows
    # print(n)

    for i in range(0, m):
        word_tokens = word_tokenize(test_dataset.cell_value(i, 1))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector1 = vectorizerStatement.transform([filtered_sentence])

        filtered_sentence = ""
        word_tokens = word_tokenize(test_dataset.cell_value(i, 2))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector2 = vectorizerStatement.transform([filtered_sentence])

        filtered_sentence = ""
        word_tokens = word_tokenize(test_dataset.cell_value(i, 3))
        for w in word_tokens:
            if w not in stop_words:
                filtered_sentence = filtered_sentence + " " + w
        vector3 = vectorizerStatement.transform([filtered_sentence])

        arr1 = vector1.toarray()
        arr2 = vector2.toarray()
        arr3 = vector3.toarray()

        tup1 = arr1.shape
        tup2 = arr2.shape
        tup3 = arr3.shape
        length1 = tup1[1]
        length2 = tup2[1]
        length3 = tup3[1]
        count1 = 0
        count2 = 0
        count3 = 0

        for j in range(0, length1):
            if arr1[0, j] != 0:
                count1 = count1 + 1
        for j in range(0, length2):
            if arr2[0, j] != 0:
                count2 = count2 + 1
        for j in range(0, length3):
            if arr3[0, j] != 0:
                count3 = count3 + 1

        sum1 = np.sum(arr1)
        sum2 = np.sum(arr2)
        sum3 = np.sum(arr3)

        if count1 != 0:
            ans1 = sum1 / count1
        else:
            ans1 = sum1
        if count2 != 0:
            ans2 = sum2 / count2
        else:
            ans2 = sum2
        if count3 != 0:
            ans3 = sum3 / count3
        else:
            ans3 = sum3

        statement.append(ans1)
        topics.append(ans2)
        speaker.append(ans3)
        verdict.append(test_dataset.cell_value(i, 0))

    for i in range(1, m):
        j = test.cell(row=i, column=1)
        j.value = verdict[i]
    workbook2.save('new_test.xlsx')

    for i in range(1, m):
        j = test.cell(row=i, column=2)
        j.value = statement[i]
    workbook2.save('new_test.xlsx')

    for i in range(1, m):
        j = test.cell(row=i, column=3)
        j.value = topics[i]
    workbook2.save('new_test.xlsx')

    for i in range(1, m):
        j = test.cell(row=i, column=4)
        j.value = speaker[i]
    workbook2.save('new_test.xlsx')

    if __name__ == "__encode_data__":
        encode_data()

# encode_data()
